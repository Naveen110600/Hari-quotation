import os
import sys
import time
import io
import webbrowser
import threading

from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from openpyxl import Workbook, load_workbook

load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXTRACTED_DIR = os.path.join(
    BASE_DIR,
    "extracted",
    "OldBilling.exe_extracted"
)

app = Flask(
    __name__,
    template_folder=os.path.join(EXTRACTED_DIR, "templates"),
    static_folder=os.path.join(EXTRACTED_DIR, "static")
)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "sri-traders-hari-app-secure-session-key-2026")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax"
)

HARI_USERNAME = os.environ.get("HARI_USERNAME", "Hari")
HARI_PASSWORD = os.environ.get("HARI_PASSWORD", "Uma@1999")

FOLDER_PATH = r"C:\SRI Traders"
FILE_PATH = os.path.join(FOLDER_PATH, "invoices.xlsx")


@app.before_request
def require_login():
    # Allow login route and static files without login
    if request.endpoint in ["login", "static"]:
        return None

    if not session.get("logged_in"):
        if request.is_json or request.path.startswith("/get-") or request.path in ["/save", "/update-invoice", "/rename-invoice", "/delete-invoice", "/pdf-action"]:
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in"):
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            error = "Please enter both username and password"
        elif username == HARI_USERNAME and password == HARI_PASSWORD:
            session["logged_in"] = True
            session["username"] = username
            return redirect(url_for("index"))
        else:
            error = "Invalid username or password"

    return render_template("login.html", error=error)


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


def ensure_excel():
    if not os.path.exists(FOLDER_PATH):
        os.makedirs(FOLDER_PATH, exist_ok=True)

    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Invoice ID", "File Name", "HTML Data"])
        wb.save(FILE_PATH)


def load_invoices():
    ensure_excel()

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    invoices = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            invoices.append({
                "id": str(row[0]),
                "name": str(row[1]) if row[1] is not None else "",
                "data": str(row[2]) if row[2] is not None else ""
            })

    return invoices


def save_invoice(file_name, html_data):
    ensure_excel()

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    invoice_id = str(int(time.time() * 1000))

    ws.append([
        invoice_id,
        file_name,
        html_data
    ])

    wb.save(FILE_PATH)

    return invoice_id


def update_invoice_excel(invoice_id, file_name, html_data):
    ensure_excel()

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and str(row[0].value) == str(invoice_id):
            row[1].value = file_name
            row[2].value = html_data
            break

    wb.save(FILE_PATH)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/get-invoices")
def get_invoices():
    invoices = load_invoices()
    return jsonify(invoices)


@app.route("/get-invoice/<invoice_id>")
def get_invoice(invoice_id):
    ensure_excel()

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None and str(row[0]) == str(invoice_id):
            return jsonify({
                "id": str(row[0]),
                "name": str(row[1]) if row[1] is not None else "",
                "data": str(row[2]) if row[2] is not None else ""
            })

    return jsonify({
        "id": str(invoice_id),
        "name": "",
        "data": ""
    })


@app.route("/save", methods=["POST"])
def save():
    data = request.json or {}

    file_name = data.get("name", "")
    html_data = data.get("data", "")

    inv_id = save_invoice(
        file_name,
        html_data
    )

    return jsonify({
        "message": "Invoice saved",
        "invoice_id": inv_id
    })


@app.route("/update-invoice", methods=["POST"])
def update_invoice():
    data = request.json or {}

    inv_id = data.get("id")
    file_name = data.get("name", "")
    html_data = data.get("data", "")

    update_invoice_excel(
        inv_id,
        file_name,
        html_data
    )

    return jsonify({
        "message": "Invoice updated"
    })


@app.route("/pdf-action", methods=["POST"])
def pdf_action():
    if "file" in request.files:
        file = request.files["file"]
        pdf_bytes = file.read()

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=False,
            download_name="invoice.pdf"
        )

    return jsonify({
        "error": "No file provided"
    }), 400


@app.route("/rename-invoice", methods=["POST"])
def rename_invoice():
    data = request.json or {}

    old_id = str(data.get("old_id", ""))
    new_id = str(data.get("new_id", ""))

    ensure_excel()

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and str(row[0].value) == old_id:
            row[1].value = new_id
            break

    wb.save(FILE_PATH)

    return jsonify({
        "message": "Renamed"
    })


@app.route("/delete-invoice", methods=["POST"])
def delete_invoice():
    data = request.json or {}

    delete_id = str(data.get("id", ""))

    ensure_excel()

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    new_rows = [rows[0]]

    for r in rows[1:]:
        if r and r[0] is not None and str(r[0]) == delete_id:
            continue

        new_rows.append(r)

    wb.remove(ws)

    ws = wb.create_sheet(title="Sheet")

    for r in new_rows:
        ws.append(r)

    wb.save(FILE_PATH)

    return jsonify({
        "message": "Deleted"
    })


# ---------------------------------------------------------
# START APPLICATION
# ---------------------------------------------------------

if __name__ == "__main__":

    def open_browser():
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe")
        ]
        opened = False
        for p in chrome_paths:
            if os.path.exists(p):
                try:
                    webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(p))
                    webbrowser.get('chrome').open("http://127.0.0.1:5000")
                    opened = True
                    break
                except Exception:
                    pass
        if not opened:
            webbrowser.open("http://127.0.0.1:5000")

    # Automatically open the billing website in Chrome
    # after the Flask server starts.
    threading.Timer(
        1.5,
        open_browser
    ).start()

    # Production-style local startup:
    # No Flask debugger
    # No debugger PIN
    # No automatic reloader
    app.run(
        host="127.0.0.1",
        port=5000,
        debug=False,
        use_reloader=False
    )
